"""XLSX parser (spec 14 T09).

Uses ``openpyxl==3.1.5`` (Spec 12 sandbox alignment per
:file:`packages/core/src/persona/sandbox/image/requirements.txt`). Gated
behind the ``[documents]`` extra (D-14-X-documents-extra); the import is
lazy inside :func:`parse_xlsx` so the package imports cleanly without the
extra. Missing-library failure surfaces as
:class:`~persona.documents.errors.MissingDependencyError`.

Extraction strategy:

- One :class:`DocumentSection` per **sheet**. The section's
  :attr:`DocumentSection.sheet` field carries the sheet name.
- Per-sheet text is a pipe-separated table (header row + data rows), with
  D-14-3's row cap applied per-sheet (default 1000; env override
  ``PERSONA_DOC_SPREADSHEET_ROW_CAP`` shared with the CSV parser).
- Oversize sheets emit a summary + first/last 50 rows + the
  :data:`FLAG_ROW_CAP_TRUNCATED` flag (the §8 tabular-boundary signal).
- :attr:`ParseResult.sheet_names` is populated with every sheet's name for
  the T16 synopsis.

Fail-safe handling:

- ``openpyxl.load_workbook`` raises ``InvalidFileException`` /
  ``BadZipFile`` / similar on corrupt files. Caught at the adapter boundary
  and re-raised as :class:`~persona.documents.errors.CorruptDocumentError`.
- A workbook with no sheets, or sheets that yield no text, raises
  ``CorruptDocumentError`` with ``reason="empty_after_decode"``.
"""

from __future__ import annotations

from typing import TYPE_CHECKING, Any

from persona.documents.chunker import DocumentSection
from persona.documents.errors import (
    CorruptDocumentError,
    MissingDependencyError,
)
from persona.documents.parsers import FLAG_ROW_CAP_TRUNCATED, ParseResult
from persona.documents.parsers.csv import SAMPLE_ROWS_PER_END, resolve_row_cap

if TYPE_CHECKING:
    from pathlib import Path

__all__ = ["parse_xlsx"]


def _import_openpyxl() -> Any:  # noqa: ANN401
    """Lazy-import ``openpyxl`` with a clear MissingDependencyError on failure."""
    try:
        import openpyxl  # noqa: PLC0415 — deliberate lazy import
    except ImportError as exc:
        raise MissingDependencyError(
            "openpyxl is not installed",
            context={
                "format": "xlsx",
                "install_hint": "pip install persona-core[documents]",
            },
        ) from exc
    return openpyxl


def parse_xlsx(path: Path) -> ParseResult:
    """Parse an ``.xlsx`` workbook to per-sheet text-table sections.

    Args:
        path: Path to the ``.xlsx`` file.

    Returns:
        :class:`ParseResult` with one :class:`DocumentSection` per sheet.
        :attr:`ParseResult.sheet_names` lists every sheet's name (used by
        the T16 synopsis). :attr:`ParseResult.flags` includes
        :data:`FLAG_ROW_CAP_TRUNCATED` when at least one sheet exceeded the
        row cap.

    Raises:
        MissingDependencyError: ``openpyxl`` is not installed.
        CorruptDocumentError: The file isn't a valid xlsx or extraction
            yielded no text (criterion #8).
    """
    openpyxl = _import_openpyxl()
    size_bytes = path.stat().st_size
    filename = path.name

    if size_bytes == 0:
        raise CorruptDocumentError(
            "file is empty",
            context={"format": "xlsx", "reason": "empty_file", "filename": filename},
        )

    try:
        workbook = openpyxl.load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — adapter-boundary catch
        reason = _classify_open_failure(exc)
        raise CorruptDocumentError(
            "could not open xlsx",
            context={"format": "xlsx", "reason": reason, "filename": filename},
        ) from exc

    sheet_names = tuple(workbook.sheetnames)
    if not sheet_names:
        raise CorruptDocumentError(
            "workbook has no sheets",
            context={
                "format": "xlsx",
                "reason": "no_sheets",
                "filename": filename,
            },
        )

    row_cap = resolve_row_cap()
    sections: list[DocumentSection] = []
    flags: list[str] = []

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        rows = list(_read_sheet_rows(sheet))
        if not rows:
            continue
        section_text, truncated = _format_sheet_table(rows, row_cap=row_cap)
        if not section_text.strip():
            continue
        if truncated and FLAG_ROW_CAP_TRUNCATED not in flags:
            flags.append(FLAG_ROW_CAP_TRUNCATED)
        sections.append(DocumentSection(text=section_text, sheet=sheet_name))

    if not sections:
        raise CorruptDocumentError(
            "extraction yielded no text",
            context={
                "format": "xlsx",
                "reason": "empty_after_decode",
                "filename": filename,
            },
        )

    return ParseResult(
        sections=tuple(sections),
        flags=tuple(flags),
        sheet_names=sheet_names,
        size_bytes=size_bytes,
    )


def _classify_open_failure(exc: BaseException) -> str:
    """Map an openpyxl open failure to a stable ``reason`` discriminator."""
    name = type(exc).__name__
    if "InvalidFile" in name:
        return "not_an_xlsx"
    if "Zip" in name or "BadZip" in name:
        return "not_a_zip"
    return "unknown"


def _read_sheet_rows(sheet: Any) -> list[list[str]]:  # noqa: ANN401
    """Read a worksheet as a list of row-lists of stringified cells.

    Trailing fully-empty rows are dropped so the row-count reported in the
    summary reflects real content. Cell ``None`` values become empty strings.
    """
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        cells = ["" if cell is None else str(cell) for cell in row]
        rows.append(cells)
    # Trim trailing empty rows.
    while rows and not any(c.strip() for c in rows[-1]):
        rows.pop()
    return rows


def _format_sheet_table(rows: list[list[str]], *, row_cap: int) -> tuple[str, bool]:
    """Render a sheet's rows as a pipe-separated text table.

    Mirrors the CSV parser's :func:`_format_table` so read-path output is
    consistent across the two tabular formats. Returns
    ``(text, truncated)``.
    """
    if not rows:
        return ("", False)

    header = rows[0]
    data_rows = rows[1:]
    total_data_rows = len(data_rows)

    if total_data_rows + 1 <= row_cap:
        formatted = " | ".join(header)
        for row in data_rows:
            formatted += "\n" + " | ".join(_normalise_row(row, width=len(header)))
        return (formatted, False)

    head = data_rows[:SAMPLE_ROWS_PER_END]
    tail = data_rows[-SAMPLE_ROWS_PER_END:]
    hidden_count = total_data_rows - len(head) - len(tail)

    lines = [
        f"[Sheet summary — {len(header)} columns, {total_data_rows} data rows]",
        " | ".join(header),
    ]
    for row in head:
        lines.append(" | ".join(_normalise_row(row, width=len(header))))
    if hidden_count > 0:
        lines.append(f"... {hidden_count} rows hidden ...")
    for row in tail:
        lines.append(" | ".join(_normalise_row(row, width=len(header))))
    lines.append(
        "(Truncated for prompt budget. Use the code-execution sandbox for "
        "computation over the full data — spec §8 tabular boundary.)"
    )
    return ("\n".join(lines), True)


def _normalise_row(row: list[str], *, width: int) -> list[str]:
    """Pad or truncate a row to ``width`` cells so the table is rectangular."""
    if len(row) < width:
        return [*row, *([""] * (width - len(row)))]
    if len(row) > width:
        return row[:width]
    return row
