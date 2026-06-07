"""Spec 16 T13 — xlsx_generation end-to-end backend production-verification test.

Quality-bar criterion #2 binary test for the ``xlsx_generation`` skill. This
is the **production-verification path** for the xlsx format: the file the
persona-runtime + persona-api wiring lands at the persona-workspace path is
opened with ``openpyxl`` and asserted against the §3.7 per-criterion test
surfaces from ``docs/specs/phase2/spec_16/research.md``.

The frontend download chip is NOT yet wired (per the user's clarification —
T13 is the backend production-verification surface; the frontend ships in a
separate spec). Backend persistence ends at
``<persona_workspace_root>/<owner_id>/<persona_id>/<filename>`` per
``D-17-X-bytes-persistence`` / ``D-12-X-read-produced-file`` — the
``_persist_produced_file`` callback the api's
:func:`make_pool_code_execution_tool` builds is the single-mechanism plumbing
this test exercises directly (without the pool/credits/contextvar machinery
the api wraps it in).

**Pre-conditions verified before composing this test:**

- ``D-16-X-6`` (PATH fix): ``LocalDockerSandbox._BASE_CONTAINER_KWARGS
  ["environment"]["PATH"]`` prepends ``/opt/venv/bin`` (verified at
  ``packages/core/src/persona/sandbox/local_docker.py:194``). The xlsx
  code snippet imports ``openpyxl`` directly — no ``sys.path.insert``
  prelude.
- ``D-16-X-7`` (supplements relative-path fix): ``collect_skill_supplements``
  emits ``.skills/<name>/supplements/<topic>.md`` relative paths (verified
  at ``packages/core/src/persona/skills/use_skill_tool.py:80``). Production
  supplements-staging works end-to-end; this test scans the real
  ``xlsx_generation`` skill with its on-disk ``supplements/`` directory so
  the use_skill intercept stages the four supplements onto the deferred
  list, and the next ``code_execution`` call receives them as
  ``input_files``.

**§3.7 surfaces exercised (programmatic; visual-only documented honestly):**

#1 ``wb["Months"]["B2"].number_format != "General"``      → programmatic
#2 ``"Months!" in wb["Summary"]["B2"].value``             → programmatic
#3 formulas-evaluate-no-#REF!/#VALUE!                     → PARTIAL surrogate
#4 ``wb["Months"].freeze_panes == "A2"``                  → programmatic
#5 column widths sized                                    → programmatic
#6 ``"Sheet1" not in wb.sheetnames and >=2 sheets``       → programmatic
#7 ``wb["Months"]["A1"].font.bold is True``               → programmatic
#8 no zero-byte computed cells                            → programmatic

Per the §3.7 row 3 note in research.md: "Load with ``data_only=True``:
``assert wb_eval["Summary"]["B2"].value not in (None, "#REF!", "#VALUE!")``
— **requires re-saving via Excel/LibreOffice first to bake values**." We
cannot bake values inside the sandbox (openpyxl writes formulas, not
results), so #3 is implemented as a **structural surrogate** that asserts
the formula string is well-formed (no literal ``#REF!`` / ``#VALUE!`` in
the formula source) and reports PARTIAL with the documented limitation.

Markers: ``integration`` + ``docker`` (mirrors T09/T10) so CI without Docker
can opt out via ``-m "not docker"``.

Inspection artifact: a copy of the produced .xlsx lands at
``docs/specs/phase2/spec_16/inspection/xlsx_sample.xlsx`` (gitignored per
``D-16-X-3``) — operator-reproducible by re-running the test.
"""

# ruff: noqa: SLF001, ANN401 — pinning the scripted backend into the tier
# registry cache (SLF001) and the MemoryStore double's loose kwargs (ANN401)
# mirror the T09/T10 conventions.

from __future__ import annotations

import shutil
import uuid
from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl
import pytest
from persona.backends import BackendConfig, ChatResponse
from persona.backends.types import TokenUsage
from persona.sandbox.local_docker import (
    DEFAULT_IMAGE,
    LocalDockerSandbox,
    is_docker_available,
)
from persona.sandbox.result import NetworkPolicy, ResourceLimits, SandboxFile
from persona.sandbox.tool import make_code_execution_tool
from persona.schema.persona import Persona, PersonaIdentity
from persona.schema.tools import ToolCall
from persona.skills import (
    SkillInjector,
    SkillScanner,
    make_use_skill_tool,
)
from persona.tools import Toolbox
from persona_runtime.agentic.loop import AgenticLoop
from persona_runtime.agentic.run import RunStatus
from persona_runtime.agentic.step import StepType
from persona_runtime.prompt import PromptBuilder
from persona_runtime.router import Router
from persona_runtime.tier import TierConfig, TierRegistry

if TYPE_CHECKING:
    from persona.backends.types import ToolSpec
    from persona.schema.conversation import ConversationMessage
    from persona.schema.skills import SkillSpec
    from persona.tools.protocol import AsyncTool


pytestmark = [pytest.mark.integration, pytest.mark.docker]


# The bundled built-in skills live alongside persona-core's source; T13 scans
# the real ``xlsx_generation`` skill so the on-disk supplements/ tree is
# staged via the M1a runtime affordance (D-16-X-7 fix verified).
_BUILTIN_ROOT = (
    Path(__file__).parent.parent.parent.parent / "core" / "src" / "persona" / "skills" / "builtin"
).resolve()

# Inspection artifact location — gitignored per D-16-X-3. The test writes a
# copy of the produced .xlsx here so the operator can reproduce evidence by
# re-running the test (per state.md scorecard discipline).
_INSPECTION_DIR = (
    Path(__file__).resolve().parents[4] / "docs" / "specs" / "phase2" / "spec_16" / "inspection"
)


_DUMMY_CFG = BackendConfig(provider="anthropic", model="m", api_key=None)


# ---------------------------------------------------------------------------
# Module-private _ScriptedBackend (T01 audit A1 — per-file api convention).
# Copied from packages/runtime/tests/_fakes.py:50; agentic chat_script mode
# only (the AgenticLoop drives chat() never chat_stream()).
# ---------------------------------------------------------------------------


class _ScriptedBackend:
    """A ChatBackend that replays a scripted sequence of ChatResponse objects.

    Mirrors :class:`persona_runtime.tests._fakes.ScriptedBackend`'s agentic
    mode. Module-private per the api per-file convention (T01 audit A1).
    """

    def __init__(
        self,
        chat_script: list[ChatResponse],
        *,
        provider_name: str = "anthropic",
        model_name: str = "claude-sonnet-4-6",
    ) -> None:
        self._chat_script = list(chat_script)
        self._chat_index = 0
        self._provider_name = provider_name
        self._model_name = model_name
        self.chat_calls = 0

    @property
    def provider_name(self) -> str:
        return self._provider_name

    @property
    def model_name(self) -> str:
        return self._model_name

    @property
    def supports_native_tools(self) -> bool:
        return False

    @property
    def supports_vision(self) -> bool:
        return False

    async def chat(
        self,
        messages: list[ConversationMessage],  # noqa: ARG002
        *,
        tools: list[ToolSpec] | None = None,  # noqa: ARG002
        temperature: float = 0.0,  # noqa: ARG002
        max_tokens: int = 4096,  # noqa: ARG002
        stop: list[str] | None = None,  # noqa: ARG002
    ) -> ChatResponse:
        self.chat_calls += 1
        if self._chat_index < len(self._chat_script):
            response = self._chat_script[self._chat_index]
            self._chat_index += 1
            return response
        # Defensive empty final — the loop should have terminated by here.
        return ChatResponse(
            content="[FINAL] ",
            tool_calls=[],
            usage=TokenUsage(prompt_tokens=1, completion_tokens=0, total_tokens=1),
            model=self._model_name,
            provider=self._provider_name,
            latency_ms=0.0,
        )

    async def chat_stream(
        self,
        messages: list[ConversationMessage],  # noqa: ARG002
        *,
        tools: list[ToolSpec] | None = None,  # noqa: ARG002
        temperature: float = 0.0,  # noqa: ARG002
        max_tokens: int = 4096,  # noqa: ARG002
        stop: list[str] | None = None,  # noqa: ARG002
    ) -> Any:
        raise NotImplementedError("AgenticLoop drives chat(), not chat_stream()")


# ---------------------------------------------------------------------------
# In-memory MemoryStore double — the loops require stores; T13 asserts on
# the produced file, not memory writes.
# ---------------------------------------------------------------------------


class _MemStore:
    def write(self, *_a: Any, **_kw: Any) -> None: ...
    def query(self, *_a: Any, **_kw: Any) -> list[Any]:
        return []

    def get_all(self, *_a: Any, **_kw: Any) -> list[Any]:
        return []

    def delete(self, *_a: Any, **_kw: Any) -> None: ...
    def remove_documents(self, *_a: Any, **_kw: Any) -> None: ...
    def history(self, *_a: Any, **_kw: Any) -> list[Any]:
        return []

    def rollback(self, *_a: Any, **_kw: Any) -> None: ...


def _empty_stores() -> dict[str, Any]:
    return {
        "identity": _MemStore(),
        "self_facts": _MemStore(),
        "worldview": _MemStore(),
        "episodic": _MemStore(),
    }


# ---------------------------------------------------------------------------
# Helpers.
# ---------------------------------------------------------------------------


def _image_present(tag: str) -> bool:
    """True if the sandbox image is locally available on the Docker daemon."""
    try:
        import docker
        from docker.errors import DockerException, ImageNotFound
    except ImportError:
        return False
    try:
        client = docker.from_env()
        try:
            client.images.get(tag)
        except ImageNotFound:
            return False
        finally:
            client.close()
    except DockerException:
        return False
    return True


def _docker_skip_reason() -> str | None:
    """Return a skip reason if Docker substrate isn't ready, else None."""
    if not is_docker_available():
        return (
            "Docker daemon unreachable; T13 needs LocalDockerSandbox to drive "
            "code_execution end-to-end. Start Docker and rerun."
        )
    if not _image_present(DEFAULT_IMAGE):
        return (
            f"Sandbox image {DEFAULT_IMAGE!r} not built. Build via the Spec 12 "
            "T06 Dockerfile (or pull) before running T13."
        )
    return None


def _persona() -> Persona:
    """Persona declaring xlsx_generation skill + code_execution + use_skill tools."""
    return Persona(
        persona_id="finance_assistant_t13",
        identity=PersonaIdentity(
            name="Bjorn",
            role="Norwegian budgeting assistant",
            background="Builds annual budgets for small businesses in NOK.",
            language_default="nb",
            constraints=["Never invent financial figures."],
        ),
        skills=["xlsx_generation"],
        tools=["use_skill", "code_execution"],
    )


def _scan_real_xlsx_skill() -> list[SkillSpec]:
    """Scan the real built-in xlsx_generation skill (with supplements on disk).

    D-16-X-7 verification: the on-disk ``supplements/`` directory at
    ``packages/core/src/persona/skills/builtin/xlsx_generation/supplements/``
    is scanned end-to-end. The use_skill intercept then calls
    ``collect_skill_supplements`` which (post-D-16-X-7) emits relative
    ``.skills/xlsx_generation/supplements/<topic>.md`` SandboxFile paths,
    which the LocalDockerSandbox seeds at
    ``/workspace/in/.skills/xlsx_generation/supplements/<topic>.md`` —
    matching the verbatim path the SKILL.md body teaches the model to read.
    """
    scanner = SkillScanner([_BUILTIN_ROOT])
    specs = scanner.scan(["xlsx_generation"])
    assert len(specs) == 1, (
        f"scanner returned {len(specs)} specs for ['xlsx_generation']; expected 1"
    )
    return specs


def _resp(content: str = "", *, tool_calls: list[ToolCall] | None = None) -> ChatResponse:
    return ChatResponse(
        content=content,
        tool_calls=tool_calls or [],
        usage=TokenUsage(prompt_tokens=10, completion_tokens=5, total_tokens=15),
        model="claude-sonnet-4-6",
        provider="anthropic",
        latency_ms=1.0,
    )


def _build_loop(
    *,
    persona: Persona,
    backend: _ScriptedBackend,
    toolbox: Toolbox,
    scanned_skills: list[SkillSpec],
    deferred_holder: list[SandboxFile],
    max_steps: int = 20,
) -> AgenticLoop:
    """Wire an :class:`AgenticLoop` over the supplied collaborators.

    Mirrors the runtime test idiom: prime the tier registry's lazy cache so
    every tier resolves to the scripted backend. Bind the shared
    ``deferred_input_files`` holder per D-16-2-state-location so the
    use_skill intercept's ``self.deferred_input_files.extend(...)`` lands in
    the same list the code_execution tool's drain-and-clear provider reads.
    """
    registry = TierRegistry(
        {
            "frontier": TierConfig(name="frontier", backend_config=_DUMMY_CFG),
            "mid": TierConfig(name="mid", backend_config=_DUMMY_CFG),
            "small": TierConfig(name="small", backend_config=_DUMMY_CFG),
        }
    )
    registry._cache = {  # type: ignore[assignment]
        "frontier": backend,  # type: ignore[dict-item]
        "mid": backend,  # type: ignore[dict-item]
        "small": backend,  # type: ignore[dict-item]
    }
    loop = AgenticLoop(
        persona=persona,
        stores=_empty_stores(),  # type: ignore[arg-type]
        toolbox=toolbox,
        skill_injector=SkillInjector(),
        scanned_skills=scanned_skills,
        prompt_builder=PromptBuilder(),
        router=Router(),
        tier_registry=registry,
        max_steps=max_steps,
    )
    # Bind the loop's public deferred_input_files attribute to the SHARED
    # holder (same list identity per D-16-2-state-location). The use_skill
    # intercept's ``self.deferred_input_files.extend(supplements)`` mutates
    # the same list the code_execution tool's provider drains.
    loop.deferred_input_files = deferred_holder
    return loop


# ---------------------------------------------------------------------------
# The representative xlsx task — §3.4 verbatim:
#
#   "Build a 2-sheet annual budget workbook. Sheet 'Months' has 12 monthly
#    columns (Jan2026 ... Dec2026) for 8 expense categories. Sheet 'Summary'
#    has one cell per category showing the year-total via a cross-sheet
#    SUMIF formula like =SUMIF(Months!$A:$A, A2, Months!$N:$N) (where N is
#    the year-total column on Months). Freeze the header row on both sheets;
#    format the totals as NOK currency."
#
# The snippet below is the minimal-but-COMPLETE openpyxl code an agent would
# write after reading the SKILL.md body — exercises the §3.7 surfaces a
# real production turn would produce. Idioms drawn directly from the
# xlsx_generation SKILL.md (cross-sheet SUMIF, freeze_panes, header font,
# number_format, descriptive sheet names, no Sheet1).
# ---------------------------------------------------------------------------


_XLSX_CODE = """\
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

wb = Workbook()
del wb["Sheet"]  # drop default to avoid the "Sheet1" trap
months = wb.create_sheet("Months")
summary = wb.create_sheet("Summary")

# --- Months sheet -----------------------------------------------------------
headers = [
    "Category", "Jan2026", "Feb2026", "Mar2026", "Apr2026", "May2026",
    "Jun2026", "Jul2026", "Aug2026", "Sep2026", "Oct2026", "Nov2026",
    "Dec2026", "Total",
]
months.append(headers)
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
for cell in months[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
months.freeze_panes = "A2"

categories = [
    "Rent", "Salaries", "Utilities", "Supplies", "Marketing",
    "Travel", "Insurance", "Software",
]
# 8 expense categories x 12 month columns of plausible NOK figures + SUM total
for row_idx, cat in enumerate(categories, start=2):
    months.cell(row=row_idx, column=1, value=cat)
    base = 1000 * row_idx  # deterministic per-row figures
    for col in range(2, 14):  # B..M = 12 months
        months.cell(row=row_idx, column=col, value=base + col * 100)
        months.cell(row=row_idx, column=col).number_format = "#,##0.00 [$NOK]"
    # Total column N — formula, not constant (per SKILL.md "always formula")
    months.cell(
        row=row_idx, column=14,
        value=f"=SUM(B{row_idx}:M{row_idx})",
    )
    months.cell(row=row_idx, column=14).number_format = "#,##0.00 [$NOK]"

# Column widths — never default 8.43 (per SKILL.md "##" trap)
months.column_dimensions["A"].width = 22  # category names wider
for col in range(2, 15):  # B..N
    months.column_dimensions[get_column_letter(col)].width = 14

# --- Summary sheet ---------------------------------------------------------
summary.append(["Category", "Year total"])
for cell in summary[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
summary.freeze_panes = "A2"

# Cross-sheet SUMIF — the §3.4 hard feature. References "Months!" by name,
# never Sheet1!. $A:$A and $N:$N are absolute; A{row_idx} is relative.
for row_idx, cat in enumerate(categories, start=2):
    summary.cell(row=row_idx, column=1, value=cat)
    summary.cell(
        row=row_idx, column=2,
        value=f"=SUMIF(Months!$A:$A, A{row_idx}, Months!$N:$N)",
    )
    summary.cell(row=row_idx, column=2).number_format = "#,##0.00 [$NOK]"

summary.column_dimensions["A"].width = 22
summary.column_dimensions["B"].width = 18

out = Path("/workspace/out/budget-2026.xlsx")
out.parent.mkdir(parents=True, exist_ok=True)
wb.save(out)
print(f"wrote {out}")
"""


# ---------------------------------------------------------------------------
# The test class.
# ---------------------------------------------------------------------------


class TestT13XlsxEndToEnd:
    """Spec 16 T13 — xlsx_generation production-verification path.

    Drives an :class:`AgenticLoop` with the real ``xlsx_generation`` skill
    scanned (so supplements are staged via M1a) + a real
    :class:`LocalDockerSandbox` + a ``_persist_produced_file`` callback that
    copies the produced .xlsx from the sandbox session's host_out into a
    persona-workspace path (mirrors the api's
    :func:`make_pool_code_execution_tool` D-17-X-bytes-persistence wiring,
    without the pool / credits / contextvar stack the api wraps).

    Opens the persona-workspace .xlsx with openpyxl and asserts each §3.7
    surface from research.md. Saves a copy to the gitignored inspection/
    directory so the operator can reproduce the evidence.
    """

    @pytest.mark.asyncio
    async def test_xlsx_generation_produces_file_meeting_quality_bar(
        self,
        tmp_path: Path,
    ) -> None:
        skip_reason = _docker_skip_reason()
        if skip_reason:
            pytest.skip(skip_reason)

        # Persona declaring xlsx_generation + the supporting tools.
        persona = _persona()
        scanned = _scan_real_xlsx_skill()
        # Sanity: the real skill must carry an on-disk supplements/ tree so
        # the D-16-X-7 production-fix verification is meaningful end-to-end.
        spec = scanned[0]
        supplements_dir = spec.path / "supplements"
        assert supplements_dir.is_dir(), (
            f"xlsx_generation skill must ship supplements/ on disk; got "
            f"{spec.path} (D-16-X-7 verification requires the real tree)."
        )
        supplement_files = list(supplements_dir.glob("*.md"))
        assert supplement_files, (
            f"supplements/ at {supplements_dir} is empty; D-16-X-7 path-fix "
            "test requires at least one supplement to stage."
        )

        # Two workspace roots:
        #   sandbox_root → LocalDockerSandbox's host_in/host_out scratch (per session)
        #   persona_workspace_root → where _persist_produced_file copies the .xlsx,
        #     mirroring the api's <workspace_root>/<owner_id>/<persona_id>/ layout.
        sandbox_root = tmp_path / "sbx"
        persona_workspace_root = tmp_path / "persona_workspaces"
        owner_id = "owner_t13_xlsx"
        # Use the persona's persona_id (post-validation, the Persona schema
        # may default to derived; the test uses the value the persona uses).
        persona_id = persona.persona_id or "finance_assistant_t13"
        persona_workspace = persona_workspace_root / owner_id / persona_id
        persona_workspace.mkdir(parents=True, exist_ok=True)

        sandbox = LocalDockerSandbox(workspace_root=sandbox_root)
        session_id = f"{owner_id}:conv_{uuid.uuid4().hex[:8]}"
        try:
            await sandbox.create_session(
                session_id, limits=ResourceLimits(), network=NetworkPolicy()
            )

            # D-16-2-state-location: shared list[SandboxFile] holder bound to
            # BOTH the loop's deferred_input_files attribute AND the
            # code_execution tool's drain-and-clear provider closure (same
            # list identity).
            deferred_holder: list[SandboxFile] = []

            def _drain_and_clear() -> list[SandboxFile]:
                snapshot = list(deferred_holder)
                deferred_holder.clear()
                return snapshot

            # D-17-X-bytes-persistence: the production callback. Copies each
            # produced file from the sandbox session's host_out into the
            # persona-workspace at <persona_workspace>/<ref>. This is the
            # single-mechanism plumbing the api's
            # make_pool_code_execution_tool wires the pool sandbox into; T13
            # exercises it directly against the LocalDockerSandbox without
            # the pool/credits/contextvar wrapper.
            async def _persist_produced_file(sid: str, ref: str) -> None:
                await sandbox.copy_produced_file_to(sid, ref, persona_workspace / ref)

            code_exec_tool: AsyncTool = make_code_execution_tool(
                sandbox,
                persona_id=persona_id,
                session_id_provider=lambda: session_id,
                deferred_input_files_provider=_drain_and_clear,
                produced_file_persister=_persist_produced_file,
            )
            use_skill_tool = make_use_skill_tool(scanned)
            toolbox = Toolbox(
                [use_skill_tool, code_exec_tool],
                allow_list=["use_skill", "code_execution"],
            )

            # Scripted backend script:
            #   (a) Activate xlsx_generation → use_skill intercept stages
            #       supplements onto the deferred_holder via M1a.
            #   (b) Run the representative openpyxl snippet via code_execution.
            #   (c) Final reply.
            chat_script: list[ChatResponse] = [
                _resp(
                    tool_calls=[
                        ToolCall(
                            name="use_skill",
                            args={"skill_name": "xlsx_generation"},
                            call_id="us1",
                        )
                    ]
                ),
                _resp(
                    tool_calls=[
                        ToolCall(
                            name="code_execution",
                            args={"code": _XLSX_CODE},
                            call_id="ce1",
                        )
                    ]
                ),
                _resp("[FINAL] Wrote /workspace/out/budget-2026.xlsx"),
            ]
            backend = _ScriptedBackend(chat_script=chat_script)
            loop = _build_loop(
                persona=persona,
                backend=backend,
                toolbox=toolbox,
                scanned_skills=scanned,
                deferred_holder=deferred_holder,
            )

            run = await loop.run(
                "Produce the 2026 annual budget workbook with monthly figures "
                "for 8 categories and a Summary sheet using cross-sheet SUMIF."
            )

            # ---- sandbox-boundary assertions -------------------------------
            assert run.status is RunStatus.COMPLETED, (
                f"expected COMPLETED; got {run.status} (error={run.error!r})"
            )
            # The code_execution step is the second TOOL_CALL (after use_skill).
            code_step = next(
                s
                for s in run.steps
                if s.type is StepType.TOOL_CALL
                and any(c.name == "code_execution" for c in s.tool_calls)
            )
            ce_result = next(
                r
                for r, c in zip(code_step.results, code_step.tool_calls, strict=True)
                if c.name == "code_execution"
            )
            assert ce_result.is_error is False, (
                f"code_execution failed: stdout/stderr={ce_result.content!r}"
            )
            assert ce_result.data is not None
            produced = ce_result.data.get("produced_files") or []
            assert any(p.get("path", "").endswith("budget-2026.xlsx") for p in produced), (
                f"expected budget-2026.xlsx in produced_files; got {produced}"
            )

            # ---- persona-workspace persistence assertion -------------------
            # D-17-X-bytes-persistence: the _persist_produced_file callback
            # copies host_out/<ref> into persona_workspace/<ref>. Acceptance
            # criterion §9 #3 (produced-files contract → workspace).
            persisted_path = persona_workspace / "budget-2026.xlsx"
            assert persisted_path.is_file(), (
                f"expected persisted .xlsx at {persisted_path}; persona workspace "
                f"contents: {list(persona_workspace.glob('*'))}"
            )
            assert persisted_path.stat().st_size > 0, (
                f"persisted .xlsx is empty at {persisted_path}"
            )

            # ---- save inspection artifact ----------------------------------
            # D-16-X-3: gitignored evidence the operator can reproduce by
            # re-running this test. mkdir parents=True for first-time runs.
            _INSPECTION_DIR.mkdir(parents=True, exist_ok=True)
            inspection_artifact = _INSPECTION_DIR / "xlsx_sample.xlsx"
            shutil.copy(persisted_path, inspection_artifact)
            assert inspection_artifact.is_file(), (
                f"inspection artifact not saved to {inspection_artifact}"
            )

            # ---- §3.7 PROGRAMMATIC ASSERTIONS ------------------------------
            # Open the persisted file with openpyxl and assert each surface.
            # All assertions use the persona-workspace copy (the production
            # artifact) — not the inspection copy.
            wb = openpyxl.load_workbook(persisted_path)

            # §3.7 #6 — Descriptive sheet names (assert BEFORE indexing so
            # an empty-sheet bug doesn't shadow the descriptive-name failure).
            assert "Sheet1" not in wb.sheetnames, (
                f"sheetnames must not contain default 'Sheet1'; got {wb.sheetnames}"
            )
            assert "Months" in wb.sheetnames, f"missing 'Months' sheet; got {wb.sheetnames}"
            assert "Summary" in wb.sheetnames, f"missing 'Summary' sheet; got {wb.sheetnames}"
            assert len(wb.sheetnames) >= 2

            months_sheet = wb["Months"]
            summary_sheet = wb["Summary"]

            # §3.7 #1 — Cell formatting applied (number_format != "General").
            # The representative task formats data cells as NOK currency.
            assert months_sheet["B2"].number_format != "General", (
                f"Months!B2 number_format should not be General; got "
                f"{months_sheet['B2'].number_format!r}"
            )

            # §3.7 #2 — Cross-sheet formula references the source sheet by name.
            # The Summary!B2 formula must contain "Months!" so a Sheet1!
            # rename trap doesn't fire (the §3.4 hard feature).
            summary_b2_formula = summary_sheet["B2"].value
            assert summary_b2_formula is not None, "Summary!B2 should carry the SUMIF formula"
            assert "Months!" in str(summary_b2_formula), (
                f"expected 'Months!' cross-sheet reference in Summary!B2; got "
                f"{summary_b2_formula!r}"
            )

            # §3.7 #3 — Formulas evaluate (no #REF!/#VALUE!).
            # PARTIAL surrogate: openpyxl writes formula strings, NOT values
            # (Excel/LibreOffice bakes values on open). The structural
            # surrogate asserts the formula string is well-formed — no
            # literal #REF! / #VALUE! in the formula source.
            assert "#REF!" not in str(summary_b2_formula), (
                f"formula string carries #REF!: {summary_b2_formula!r}"
            )
            assert "#VALUE!" not in str(summary_b2_formula), (
                f"formula string carries #VALUE!: {summary_b2_formula!r}"
            )

            # §3.7 #4 — Freeze pane on header row.
            assert months_sheet.freeze_panes == "A2", (
                f"Months sheet freeze_panes should be 'A2'; got {months_sheet.freeze_panes!r}"
            )

            # §3.7 #5 — Column widths sized (no default ~8.43).
            assert months_sheet.column_dimensions["A"].width > 8, (
                f"Months!A column width should be > 8; got "
                f"{months_sheet.column_dimensions['A'].width}"
            )

            # §3.7 #7 — Header row styled (bold).
            # Research surface checks BOTH bold and fill colour; some openpyxl
            # versions render the fill rgb as a Theme object instead of a
            # hex string when PatternFill colour was set by theme index. We
            # assert the bold (load-bearing for the "styled header" claim)
            # and check the fill exists (fill.fill_type set means a fill is
            # applied — the strongest cross-version-stable surrogate).
            assert months_sheet["A1"].font.bold is True, (
                "Months!A1 header should be bold per the styled-header surface"
            )
            assert months_sheet["A1"].fill.fill_type is not None, (
                f"Months!A1 should have a fill applied; got {months_sheet['A1'].fill.fill_type!r}"
            )

            # §3.7 #8 — No zero-byte cells in computed rows.
            # The year-total column N is row-formulated for each category;
            # every category row 2..9 should have a non-None value.
            for r in range(2, 10):
                cell = months_sheet.cell(row=r, column=14)
                assert cell.value is not None, (
                    f"Months!N{r} (year-total column) should carry a formula; got None"
                )

            wb.close()

        finally:
            await sandbox.aclose()
